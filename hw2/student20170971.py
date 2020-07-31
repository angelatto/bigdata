#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
sheet = wb['Sheet1']

total=0
i = 2
list = []
score = []
while sheet.cell(row = i, column = 1).value is not None:
    midterm = sheet.cell(row = i, column = 3).value
    final = sheet.cell(row = i, column = 4).value
    hw = sheet.cell(row = i, column = 5).value
    attendance = sheet.cell(row = i, column = 6).value
    total = midterm*0.3 + final*0.35 + hw*0.34 + attendance
    sheet.cell(row = i, column = 7, value = "{}".format(total))
    list.append(total)
    score.append(total)
    i += 1
    
    
score.sort(reverse=True)

maxA = int(len(score)*0.3)
maxB = int(len(score)*0.7)

grade = []
for i in range(0, len(score)):
    if i < maxA:
        grade.append('A0')
    elif i < maxB:
        grade.append('B0')
    else:
        grade.append('C0')


student = zip(score, grade)
dic = dict(student)

keylist = dic.keys()
for key in keylist:
    if key < 40:
        dic[key] = 'F' 


countA = 0
countB = 0
countC = 0

valuelist = dic.values()
for value in valuelist:
    if value == 'A0':
        countA += 1
    elif value == 'B0':
        countB += 1
    elif value == 'C0':
        countC += 1    


i = 0
for key in keylist: 
    if i < int(countA * 0.5):
       dic[key] = 'A+'
    elif i >= countA and i < countA + int(countB * 0.5):
       dic[key] = 'B+'
    elif i >= countB and i < countB + int(countC * 0.5):
       dic[key] = 'C+'
    i += 1


i = 2
while sheet.cell(row = i, column = 1).value is not None:
    g = float(sheet.cell(row = i, column = 7).value)
    sheet.cell(row = i, column = 8, value = dic[g])
    i += 1
        

wb.save('student.xlsx')
